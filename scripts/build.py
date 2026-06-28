#!/usr/bin/env python3
"""EVI Dashboard Builder v2 — Run: python3 scripts/build.py"""
import json, datetime, statistics, zipfile, re, sys, io
from pathlib import Path

ROOT   = Path(__file__).parent.parent
EXCEL  = ROOT / "data" / "EVI_2025-26.xlsx"
OUTPUT = ROOT / "docs" / "index.html"
WINDOW = 252  # ~1 trading year for YoY EPS growth

def load_workbook_safe(path):
    from openpyxl import load_workbook
    with zipfile.ZipFile(path, "r") as z:
        files = {n: z.read(n) for n in z.namelist()}
    styles = files.get("xl/styles.xml", b"")
    styles = re.sub(
        rb'(<family val=")(\d+)(")',
        lambda m: m.group(0) if int(m.group(2)) <= 14 else m.group(1) + b"2" + m.group(3),
        styles,
    )
    files["xl/styles.xml"] = styles
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for name, data in files.items():
            z.writestr(name, data)
    buf.seek(0)
    return load_workbook(buf, data_only=True)

def extract_data(wb):
    ws = wb["EVI 2025"]
    rows = []
    for r in range(2, ws.max_row + 1):
        def v(col):
            val = ws.cell(row=r, column=col).value
            if val is None: return 0.0
            if isinstance(val, (datetime.datetime, datetime.date)): return val
            try: return float(val)
            except: return 0.0
        date_raw = ws.cell(row=r, column=2).value
        if date_raw is None: continue
        if isinstance(date_raw, (datetime.datetime, datetime.date)):
            dt = date_raw.date() if isinstance(date_raw, datetime.datetime) else date_raw
        else:
            try: dt = datetime.date(1899, 12, 30) + datetime.timedelta(days=int(float(date_raw)))
            except: continue
        nifty, pe = v(12), v(15)
        if nifty == 0 or pe == 0: continue
        mc_inr  = v(3)   # Market cap INR crores
        usdinr  = v(5)
        mid_pe  = v(22)
        sc_pe   = v(26)
        rows.append({
            "date":               dt.isoformat(),
            "nifty50":            nifty,
            "pe":                 pe,
            "pb":                 v(13),
            "eps":                v(10) if v(10) > 0 else round(nifty / pe, 4),
            "earning_yield":      v(9),
            "india_10yr":         v(11),
            "us_10yr":            v(16),
            "yield_gap":          v(14),
            "usdinr":             usdinr,
            "dollar_index":       v(18),
            "marketcap_inr":      mc_inr,
            "marketcap_trillion": round(mc_inr * 1e7 / usdinr / 1e12, 2) if usdinr > 0 else 0,
            "marketcap_gdp":      v(7),
            "beer":               v(8),
            "t91_raw":            v(20),
            "preity":             round(v(15) * v(20), 4) if v(15) and v(20) else 0,
            "midcap150":          v(21),
            "midcap_pe":          mid_pe,
            "midcap_eps":         round(v(21) / mid_pe, 2) if mid_pe > 0 else 0,
            "midcap_earn_yield":  v(24),
            "smallcap250":        v(25),
            "smallcap_pe":        sc_pe,
            "smallcap_eps":       round(v(25) / sc_pe, 2) if sc_pe > 0 else 0,
            "smallcap_earn_yield":v(28),
        })

    # Calculate YoY EPS growth (252 trading days back) for all three indices
    for i, r in enumerate(rows):
        if i >= WINDOW:
            prev = rows[i - WINDOW]
            r["nifty_eps_growth"]    = round((r["eps"]          - prev["eps"])          / prev["eps"]          * 100, 2) if prev["eps"]          > 0 else 0
            r["midcap_eps_growth"]   = round((r["midcap_eps"]   - prev["midcap_eps"])   / prev["midcap_eps"]   * 100, 2) if prev["midcap_eps"]   > 0 else 0
            r["smallcap_eps_growth"] = round((r["smallcap_eps"] - prev["smallcap_eps"]) / prev["smallcap_eps"] * 100, 2) if prev["smallcap_eps"] > 0 else 0
        else:
            r["nifty_eps_growth"] = r["midcap_eps_growth"] = r["smallcap_eps_growth"] = 0
    return rows

def build_chart_data(rows):
    n = len(rows)
    thin_idx = list(range(0, max(0, n - 90), 3)) + list(range(max(0, n - 90), n))
    def pick(key):
        return [round(rows[i][key], 4) if isinstance(rows[i][key], float) else rows[i][key] for i in thin_idx]
    keys = ["date","nifty50","midcap150","smallcap250","pe","pb","earning_yield","india_10yr","us_10yr","yield_gap",
            "usdinr","dollar_index","marketcap_gdp","marketcap_trillion","beer","preity",
            "midcap_earn_yield","smallcap_earn_yield",
            "eps","midcap_eps","smallcap_eps",
            "nifty_eps_growth","midcap_eps_growth","smallcap_eps_growth"]
    return {k: pick(k) for k in keys}

def compute_champion(rows):
    """Backtest-derived Champion Signal: P/E + MarketCap/GDP composite.
    Returns current score (0=cheap,100=expensive), bucket, expected 12M return,
    and the historical quintile table."""
    import statistics as _st
    n = len(rows)
    pe    = [r["pe"] for r in rows]
    mcgdp = [r["marketcap_gdp"] if r["marketcap_gdp"] > 0 else None for r in rows]
    nifty = [r["nifty50"] for r in rows]

    # Forward 12M (252 trading day) returns
    W = 252
    fwd = [None]*n
    for i in range(n - W):
        if nifty[i] > 0:
            fwd[i] = (nifty[i+W] - nifty[i]) / nifty[i] * 100

    # Percentile rank helper (expensiveness: high value = high pct)
    def pctile_series(arr):
        valid = [x for x in arr if x is not None]
        out = []
        for x in arr:
            if x is None:
                out.append(None)
            else:
                out.append(sum(1 for y in valid if y <= x) / len(valid) * 100)
        return out

    pe_pct = pctile_series(pe)
    mc_pct = pctile_series(mcgdp)

    # Champion score = average of the two percentiles (skip missing)
    champ = []
    for i in range(n):
        parts = [p for p in (pe_pct[i], mc_pct[i]) if p is not None]
        champ.append(sum(parts)/len(parts) if parts else None)

    # Quintile buckets vs forward 12M return
    pairs = [(champ[i], fwd[i]) for i in range(n) if champ[i] is not None and fwd[i] is not None]
    pairs.sort(key=lambda x: x[0])
    buckets = []
    if pairs:
        cs = [p[0] for p in pairs]
        edges = [_st.quantiles(cs, n=5)[k] for k in range(4)] if len(cs) >= 5 else []
        def bucket_of(score):
            if not edges: return 0
            for k, e in enumerate(edges):
                if score < e: return k
            return 4
        grp = {k: [] for k in range(5)}
        for score, ret in pairs:
            grp[bucket_of(score)].append(ret)
        buckets = [round(sum(grp[k])/len(grp[k]), 1) if grp[k] else 0 for k in range(5)]
    else:
        edges = []

    # Current reading (last non-null champ)
    cur_score = next((champ[i] for i in range(n-1, -1, -1) if champ[i] is not None), 50)
    def bucket_of2(score):
        if not edges: return 2
        for k, e in enumerate(edges):
            if score < e: return k
        return 4
    cur_bucket = bucket_of2(cur_score)
    exp_return = buckets[cur_bucket] if buckets else 0

    # Hit rates
    cheap = [r for s, r in pairs if s < 40]
    exp   = [r for s, r in pairs if s > 60]
    cheap_hit = round(sum(1 for x in cheap if x > 0)/len(cheap)*100) if cheap else 0
    cheap_avg = round(sum(cheap)/len(cheap), 1) if cheap else 0
    exp_hit   = round(sum(1 for x in exp if x > 0)/len(exp)*100) if exp else 0
    exp_avg   = round(sum(exp)/len(exp), 1) if exp else 0

    label = "CHEAP" if cur_score < 40 else "EXPENSIVE" if cur_score > 60 else "FAIR"
    return {
        "score": round(cur_score, 1),
        "bucket": cur_bucket,
        "label": label,
        "exp_return": exp_return,
        "buckets": buckets,
        "cheap_hit": cheap_hit, "cheap_avg": cheap_avg,
        "exp_hit": exp_hit, "exp_avg": exp_avg,
        "pe_pctile": round(next((pe_pct[i] for i in range(n-1,-1,-1) if pe_pct[i] is not None), 50)),
        "mc_pctile": round(next((mc_pct[i] for i in range(n-1,-1,-1) if mc_pct[i] is not None), 50)),
    }

def compute_stats(rows):
    def med(key):
        vals = [r[key] for r in rows if r[key] > 0]
        return round(statistics.median(vals), 4) if vals else 0
    latest = rows[-1]
    return {
        "last_date": latest["date"], "total_rows": len(rows), "date_from": rows[0]["date"],
        "nifty": latest["nifty50"], "pe": latest["pe"], "pb": latest["pb"],
        "earning_yield": latest["earning_yield"], "india_10yr": latest["india_10yr"],
        "us_10yr": latest["us_10yr"], "yield_gap": latest["yield_gap"],
        "usdinr": latest["usdinr"], "dollar_index": latest["dollar_index"],
        "beer": latest["beer"], "marketcap_gdp": latest["marketcap_gdp"],
        "marketcap_trillion": latest["marketcap_trillion"],
        "preity": latest["preity"],
        "midcap_earn_yield": latest["midcap_earn_yield"],
        "smallcap_earn_yield": latest["smallcap_earn_yield"],
        "nifty_eps": latest["eps"],
        "nifty_eps_growth": latest["nifty_eps_growth"],
        "midcap_eps": latest["midcap_eps"],
        "midcap_eps_growth": latest["midcap_eps_growth"],
        "smallcap_eps": latest["smallcap_eps"],
        "smallcap_eps_growth": latest["smallcap_eps_growth"],
        "pe_median": med("pe"), "beer_median": med("beer"),
        "mcgdp_median": med("marketcap_gdp"), "yg_median": med("yield_gap"),
    }

HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>EVI Dashboard — __LAST_DATE__</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
@import url('https://fonts.googleapis.com/css2?family=Newsreader:opsz,wght@6..72,400;6..72,500;6..72,600;6..72,700&family=IBM+Plex+Mono:wght@400;500;600&family=IBM+Plex+Sans:wght@400;500;600&display=swap');
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{
  --bg:#f3efe4;--bg2:#fffefb;--bg3:#ece6d7;--bg4:#e3dcc9;
  --border:#dcd3bf;--border2:#c8bda3;
  --green:#1d4634;--green2:#2f6149;--green-dim:#5a7d6c;--green-deep:#13302339;
  --brass:#9a7321;--brass2:#b8902f;--brass-dim:#c9b787;
  --pos:#2c6e49;--neg:#a13c2b;--amber:#a8761f;--plum:#6b4a6e;--teal:#2a6f6b;
  --text:#374a40;--text2:#6b7a6f;--muted:#9aa499;--ink:#15281e;
  --mono:'IBM Plex Mono',monospace;--sans:'IBM Plex Sans',sans-serif;--serif:'Newsreader',Georgia,serif;
}
body{font-family:var(--sans);background:var(--bg);color:var(--text);
  background-image:radial-gradient(ellipse 70% 40% at 50% -5%,rgba(154,115,33,0.06),transparent);}

/* HEADER */
.hdr{background:linear-gradient(180deg,#fffefb,#f8f4ea);border-bottom:1px solid var(--border);
  padding:15px 30px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:100;
  box-shadow:0 1px 0 rgba(154,115,33,0.18),0 2px 12px rgba(21,40,30,0.04)}
.hdr-t{font-family:var(--serif);font-size:22px;font-weight:500;color:var(--ink);letter-spacing:.01em;font-style:italic}
.hdr-t b{color:var(--green);font-style:normal;font-weight:700}
.hdr-s{font-family:var(--mono);font-size:9px;color:var(--green-dim);margin-left:14px;letter-spacing:.22em;text-transform:uppercase;border-left:1px solid var(--border2);padding-left:14px}
.hdr-r{font-family:var(--mono);font-size:10px;color:var(--text2);text-align:right;letter-spacing:.04em}
.hdr-r strong{color:var(--brass);font-weight:600}
.wrap{max-width:1480px;margin:0 auto;padding:18px 28px 40px}

/* SECTION HEADERS */
.sec{font-family:var(--mono);font-size:9px;font-weight:600;letter-spacing:.28em;text-transform:uppercase;
  color:var(--green-dim);margin:22px 0 12px;display:flex;align-items:center;gap:12px}
.sec::before{content:'';width:5px;height:5px;background:var(--brass);transform:rotate(45deg);flex-shrink:0}
.sec::after{content:'';flex:1;height:1px;background:linear-gradient(90deg,var(--border2),transparent)}
.sec span{color:var(--brass)}

/* EVI COMPOSITE */
.evi-row{display:flex;gap:22px;align-items:center;background:linear-gradient(135deg,#fffefb,#faf6ec);
  border:1px solid var(--border);border-radius:10px;padding:14px 22px;margin-bottom:12px;position:relative;overflow:hidden;
  box-shadow:0 1px 3px rgba(21,40,30,0.04)}
.evi-row::before{content:'';position:absolute;left:0;top:0;bottom:0;width:3px;background:linear-gradient(180deg,var(--brass2),var(--brass))}
.evi-num{font-family:var(--serif);font-size:46px;font-weight:600;line-height:1;color:var(--green);min-width:78px;font-feature-settings:'tnum'}
.evi-txt{flex:1}
.evi-hl{font-family:var(--serif);font-size:17px;font-weight:600;color:var(--ink);margin-bottom:3px}
.evi-desc{font-size:11px;color:var(--text2);line-height:1.5;max-width:560px}
.evi-bar-wrap{min-width:240px}
.evi-bar-bg{height:6px;background:var(--bg4);border-radius:3px;margin-bottom:6px;position:relative}
.evi-bar-fill{height:100%;border-radius:3px;background:linear-gradient(90deg,var(--pos),var(--amber),var(--neg));opacity:.9}
.evi-dot{position:absolute;top:-4px;width:14px;height:14px;background:#fffefb;border-radius:50%;border:2px solid var(--green);transform:translateX(-50%);box-shadow:0 0 0 1px var(--brass),0 2px 6px rgba(21,40,30,0.2)}
.evi-bar-lbl{display:flex;justify-content:space-between;font-family:var(--mono);font-size:8px;color:var(--muted);letter-spacing:.1em;text-transform:uppercase}

/* CHAMPION SIGNAL — hero */
.champ-row{display:grid;grid-template-columns:1.45fr 1fr;gap:12px;margin-bottom:12px}
.champ-card{background:
    radial-gradient(circle at 0% 0%,rgba(154,115,33,0.07),transparent 45%),
    linear-gradient(135deg,#fffefb,#f9f5ea);
  border:1px solid var(--border);border-radius:10px;padding:18px 22px;position:relative;
  box-shadow:0 2px 8px rgba(21,40,30,0.05)}
.champ-card::after{content:'';position:absolute;inset:0;border-radius:10px;padding:1px;
  background:linear-gradient(135deg,rgba(154,115,33,0.4),transparent 42%);
  -webkit-mask:linear-gradient(#000 0 0) content-box,linear-gradient(#000 0 0);
  -webkit-mask-composite:xor;mask-composite:exclude;pointer-events:none}
.champ-head{display:flex;align-items:center;gap:10px;margin-bottom:14px}
.champ-badge{font-family:var(--mono);font-size:8px;font-weight:600;letter-spacing:.16em;text-transform:uppercase;
  color:#fffefb;background:linear-gradient(135deg,var(--green2),var(--green));padding:3px 9px;border-radius:3px;
  box-shadow:0 1px 4px rgba(29,70,52,0.3)}
.champ-title{font-family:var(--mono);font-size:9px;font-weight:500;letter-spacing:.14em;text-transform:uppercase;color:var(--green-dim)}
.champ-main{display:flex;align-items:center;gap:24px}
.champ-score{font-family:var(--serif);font-size:60px;font-weight:700;line-height:.9;min-width:96px;font-feature-settings:'tnum';
  background:linear-gradient(160deg,var(--green2),var(--green));-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text}
.champ-detail{flex:1}
.champ-label{font-family:var(--serif);font-size:16px;font-weight:600;margin-bottom:4px;letter-spacing:.01em}
.champ-sub{font-size:11px;color:var(--text2);line-height:1.5}
.champ-sub b{color:var(--ink);font-family:var(--mono);font-size:10px}
.champ-exp{text-align:center;padding:10px 16px;background:rgba(44,110,73,0.07);border:1px solid rgba(44,110,73,0.2);border-radius:7px}
.champ-exp-v{font-family:var(--serif);font-size:26px;font-weight:700;color:var(--green);line-height:1;font-feature-settings:'tnum'}
.champ-exp-l{font-family:var(--mono);font-size:8px;color:var(--muted);margin-top:4px;text-transform:uppercase;letter-spacing:.12em}
.champ-buckets{display:flex;gap:5px;margin-top:12px}
.cb{flex:1;text-align:center;padding:7px 2px;border-radius:5px;background:var(--bg3);border:1px solid transparent;transition:transform .15s}
.cb.active{border-color:var(--brass);background:rgba(154,115,33,0.1);transform:translateY(-2px)}
.cb-v{font-family:var(--mono);font-size:12px;font-weight:600}
.cb-l{font-family:var(--mono);font-size:7px;color:var(--muted);margin-top:2px;text-transform:uppercase;letter-spacing:.06em}
.champ-hits{display:flex;flex-direction:column;justify-content:center;gap:14px;height:100%}
.hit-row{display:flex;align-items:center;gap:12px}
.hit-ball{width:46px;height:46px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-family:var(--serif);font-size:15px;font-weight:700;flex-shrink:0;border:1.5px solid currentColor}
.hit-txt{font-size:11px;color:var(--text2);line-height:1.45}
.hit-txt b{color:var(--ink);font-family:var(--mono);font-size:11px}

/* KPI STRIP */
.kpi-strip{display:grid;grid-template-columns:repeat(8,1fr);gap:1px;background:var(--border);
  border:1px solid var(--border);border-radius:8px;overflow:hidden;margin-bottom:10px;box-shadow:0 1px 3px rgba(21,40,30,0.04)}
.kpi{background:var(--bg2);padding:10px 12px;transition:background .15s}
.kpi:hover{background:#fbf8f0}
.kpi-l{font-family:var(--mono);font-size:8px;font-weight:500;letter-spacing:.1em;text-transform:uppercase;color:var(--muted);margin-bottom:5px}
.kpi-v{font-family:var(--mono);font-size:16px;font-weight:600;color:var(--ink);line-height:1;font-feature-settings:'tnum'}
.kpi-v.g{color:var(--pos)}.kpi-v.r{color:var(--neg)}.kpi-v.gld{color:var(--brass)}.kpi-v.ac{color:var(--teal)}.kpi-v.pu{color:var(--plum)}
.kpi-c{font-family:var(--mono);font-size:9px;color:var(--muted);margin-top:3px}
.kpi-c.up{color:var(--pos)}.kpi-c.dn{color:var(--neg)}

/* FILTER */
.filter-row{display:flex;gap:6px;margin:14px 0 4px;align-items:center}
.filter-row::before{content:'Range';font-family:var(--mono);font-size:8px;letter-spacing:.16em;text-transform:uppercase;color:var(--muted);margin-right:6px}
.fb{font-family:var(--mono);font-size:9px;font-weight:500;letter-spacing:.08em;padding:5px 13px;border-radius:4px;border:1px solid var(--border2);background:transparent;color:var(--text2);cursor:pointer;transition:all .15s}
.fb:hover{border-color:var(--green);color:var(--green)}
.fb.on{background:linear-gradient(135deg,var(--green2),var(--green));border-color:var(--green);color:#fffefb;font-weight:600}

/* GAUGES */
.gauge-row{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:10px}
.gauge-card{background:linear-gradient(135deg,#fffefb,#faf6ec);border:1px solid var(--border);border-radius:9px;padding:13px 15px;box-shadow:0 1px 3px rgba(21,40,30,0.04)}
.gauge-title{font-family:var(--mono);font-size:8px;font-weight:600;letter-spacing:.12em;text-transform:uppercase;color:var(--green-dim);margin-bottom:10px}
.gw{display:flex;align-items:center;gap:12px}
.g-arc{position:relative;width:58px;height:58px;flex-shrink:0}
.g-arc svg{width:100%;height:100%}
.g-num{position:absolute;inset:0;display:flex;flex-direction:column;align-items:center;justify-content:center;font-family:var(--mono);font-size:12px;font-weight:600;color:var(--ink);line-height:1}
.g-pct{font-size:7px;color:var(--muted);margin-top:2px;letter-spacing:.08em}
.g-val{font-family:var(--serif);font-size:20px;font-weight:600;color:var(--ink);font-feature-settings:'tnum'}
.g-med{font-family:var(--mono);font-size:9px;color:var(--muted);margin-top:2px}
.g-zone{display:inline-block;margin-top:6px;padding:2px 7px;border-radius:3px;font-family:var(--mono);font-size:8px;font-weight:600;letter-spacing:.06em;text-transform:uppercase}
.zc{background:rgba(44,110,73,.14);color:var(--pos)}.zf{background:rgba(168,118,31,.16);color:var(--amber)}.zr{background:rgba(161,60,43,.14);color:var(--neg)}

/* CHARTS */
.g2{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:10px}
.g3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;margin-bottom:10px}
.g4{display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:10px;margin-bottom:10px}
.cc{background:linear-gradient(135deg,#fffefb,#faf6ec);border:1px solid var(--border);border-radius:9px;padding:12px 14px;transition:border-color .15s;box-shadow:0 1px 3px rgba(21,40,30,0.04)}
.cc:hover{border-color:var(--border2)}
.cc.sp2{grid-column:span 2}.cc.sp3{grid-column:span 3}
.ch{display:flex;justify-content:space-between;align-items:baseline;margin-bottom:8px}
.ct{font-family:var(--mono);font-size:9px;font-weight:600;letter-spacing:.1em;text-transform:uppercase;color:var(--text)}
.cv{font-family:var(--serif);font-size:15px;font-weight:600;color:var(--green);font-feature-settings:'tnum'}
.cw{height:135px;position:relative}
.cw canvas{display:block}
.lgd{display:flex;gap:14px;margin-top:6px;flex-wrap:wrap}
.li{display:flex;align-items:center;gap:5px;font-family:var(--mono);font-size:8px;color:var(--muted);letter-spacing:.04em}
.ld{width:7px;height:7px;border-radius:50%}

@media(max-width:980px){.kpi-strip{grid-template-columns:repeat(4,1fr)}.gauge-row,.g3,.g4{grid-template-columns:repeat(2,1fr)}.g2,.champ-row{grid-template-columns:1fr}.cc.sp2,.cc.sp3{grid-column:span 1}.champ-main{flex-wrap:wrap}}
@media(max-width:560px){.kpi-strip,.gauge-row,.g3,.g4{grid-template-columns:1fr}.hdr{flex-direction:column;gap:8px;align-items:flex-start}.hdr-r{text-align:left}}
</style>
</head>
<body>
<div class="hdr">
  <div style="display:flex;align-items:baseline">
    <span class="hdr-t"><b>EVI</b> Equity Valuation Index</span>
    <span class="hdr-s">India Market Monitor</span>
  </div>
  <div class="hdr-r">As of <strong>__LAST_DATE__</strong><br>__DATE_FROM__ &ndash; __LAST_DATE__ &middot; __TOTAL_ROWS__ sessions</div>
</div>

<div class="wrap">

  <!-- EVI COMPOSITE -->
  <div class="evi-row">
    <div class="evi-num" id="eviNum">—</div>
    <div class="evi-txt"><div class="evi-hl" id="eviHl">—</div><div class="evi-desc" id="eviDesc">—</div></div>
    <div class="evi-bar-wrap">
      <div class="evi-bar-bg"><div class="evi-bar-fill" style="width:100%"></div><div class="evi-dot" id="eviDot"></div></div>
      <div class="evi-bar-lbl"><span>Cheap</span><span>Fair</span><span>Expensive</span></div>
    </div>
  </div>

  <!-- CHAMPION SIGNAL -->
  <div class="champ-row">
    <div class="champ-card">
      <div class="champ-head">
        <span class="champ-badge">★ CFA Champion</span>
        <span class="champ-title">Verified Valuation Signal — P/E + MarketCap/GDP</span>
      </div>
      <div class="champ-main">
        <div class="champ-score" id="champScore">—</div>
        <div class="champ-detail">
          <div class="champ-label" id="champLabel">—</div>
          <div class="champ-sub" id="champSub">—</div>
          <div class="champ-buckets" id="champBuckets"></div>
        </div>
        <div class="champ-exp">
          <div class="champ-exp-v" id="champExp">—</div>
          <div class="champ-exp-l">Expected 12M return</div>
        </div>
      </div>
    </div>
    <div class="champ-card">
      <div class="champ-head"><span class="champ-title">Historical Hit Rate (12M forward)</span></div>
      <div class="champ-hits">
        <div class="hit-row">
          <div class="hit-ball" id="hitCheapBall" style="background:rgba(0,214,143,.18);color:var(--green)">—</div>
          <div class="hit-txt">When signal reads <b>CHEAP</b> (&lt;40):<br>positive <b id="hitCheapPct">—</b> of the time · avg <b id="hitCheapAvg">—</b></div>
        </div>
        <div class="hit-row">
          <div class="hit-ball" id="hitExpBall" style="background:rgba(255,92,106,.18);color:var(--red)">—</div>
          <div class="hit-txt">When signal reads <b>EXPENSIVE</b> (&gt;60):<br>positive <b id="hitExpPct">—</b> of the time · avg <b id="hitExpAvg">—</b></div>
        </div>
      </div>
    </div>
  </div>

  <!-- KPI ROW 1: Core -->
  <div class="kpi-strip">
    <div class="kpi"><div class="kpi-l">Nifty 50</div><div class="kpi-v" id="kN">—</div><div class="kpi-c" id="kNc">—</div></div>
    <div class="kpi"><div class="kpi-l">P/E</div><div class="kpi-v" id="kPE">—</div><div class="kpi-c" id="kPEm">—</div></div>
    <div class="kpi"><div class="kpi-l">P/B</div><div class="kpi-v" id="kPB">—</div><div class="kpi-c">price/book</div></div>
    <div class="kpi"><div class="kpi-l">Earning Yield</div><div class="kpi-v g" id="kEY">—</div><div class="kpi-c">1/PE×100</div></div>
    <div class="kpi"><div class="kpi-l">India 10yr</div><div class="kpi-v gld" id="kI10">—</div><div class="kpi-c" id="kYG">—</div></div>
    <div class="kpi"><div class="kpi-l">US 10yr</div><div class="kpi-v" id="kU10">—</div><div class="kpi-c" id="kSp">—</div></div>
    <div class="kpi"><div class="kpi-l">USD/INR</div><div class="kpi-v" id="kFX">—</div><div class="kpi-c" id="kDXY">—</div></div>
    <div class="kpi"><div class="kpi-l">MC/GDP</div><div class="kpi-v" id="kMC">—</div><div class="kpi-c">Buffett</div></div>
  </div>

  <!-- KPI ROW 2: Extended -->
  <div class="kpi-strip" style="margin-bottom:10px">
    <div class="kpi"><div class="kpi-l">Market Cap</div><div class="kpi-v ac" id="kMCT">—</div><div class="kpi-c">USD Trillion</div></div>
    <div class="kpi"><div class="kpi-l">PREITY Ratio</div><div class="kpi-v pu" id="kPREITY">—</div><div class="kpi-c">Nifty/US10yr</div></div>
    <div class="kpi"><div class="kpi-l">Midcap EY</div><div class="kpi-v g" id="kMEY">—</div><div class="kpi-c">Midcap 150</div></div>
    <div class="kpi"><div class="kpi-l">Smallcap EY</div><div class="kpi-v g" id="kSEY">—</div><div class="kpi-c">SC 250</div></div>
    <div class="kpi"><div class="kpi-l">Nifty EPS Growth</div><div class="kpi-v" id="kNEG">—</div><div class="kpi-c">YoY ~1yr</div></div>
    <div class="kpi"><div class="kpi-l">Midcap EPS</div><div class="kpi-v" id="kMEG">—</div><div class="kpi-c">per unit</div></div>
    <div class="kpi"><div class="kpi-l">Smallcap EPS</div><div class="kpi-v" id="kSEG">—</div><div class="kpi-c">per unit</div></div>
    <div class="kpi"><div class="kpi-l">BEER Ratio</div><div class="kpi-v gld" id="kBEER">—</div><div class="kpi-c">EY/Bond Yield</div></div>
  </div>

  <!-- FILTER -->
  <div class="filter-row" id="fr">
    <button class="fb" data-r="90">3M</button>
    <button class="fb" data-r="180">6M</button>
    <button class="fb on" data-r="365">1Y</button>
    <button class="fb" data-r="730">2Y</button>
    <button class="fb" data-r="9999">All</button>
  </div>

  <!-- GAUGES -->
  <div class="sec">Valuation Gauges — Percentile Rank vs Full History</div>
  <div class="gauge-row">
    <div class="gauge-card"><div class="gauge-title">P/E Ratio</div><div class="gw"><div class="g-arc"><svg viewBox="0 0 56 56"><circle cx="28" cy="28" r="22" fill="none" stroke="#e3dcc9" stroke-width="6" stroke-dasharray="138.2" stroke-linecap="round" transform="rotate(-90 28 28)"/><circle cx="28" cy="28" r="22" fill="none" id="arcPE" stroke="#1d4634" stroke-width="6" stroke-dasharray="138.2" stroke-dashoffset="138.2" stroke-linecap="round" transform="rotate(-90 28 28)"/></svg><div class="g-num"><span id="pPE">—</span><span class="g-pct">%ile</span></div></div><div><div class="g-val" id="vPE">—</div><div class="g-med">Med: __PE_MED__</div><div class="g-zone" id="zPE">—</div></div></div></div>
    <div class="gauge-card"><div class="gauge-title">BEER Ratio</div><div class="gw"><div class="g-arc"><svg viewBox="0 0 56 56"><circle cx="28" cy="28" r="22" fill="none" stroke="#e3dcc9" stroke-width="6" stroke-dasharray="138.2" stroke-linecap="round" transform="rotate(-90 28 28)"/><circle cx="28" cy="28" r="22" fill="none" id="arcBEER" stroke="#9a7321" stroke-width="6" stroke-dasharray="138.2" stroke-dashoffset="138.2" stroke-linecap="round" transform="rotate(-90 28 28)"/></svg><div class="g-num"><span id="pBEER">—</span><span class="g-pct">%ile</span></div></div><div><div class="g-val" id="vBEER">—</div><div class="g-med">Med: __BEER_MED__</div><div class="g-zone" id="zBEER">—</div></div></div></div>
    <div class="gauge-card"><div class="gauge-title">MC / GDP %</div><div class="gw"><div class="g-arc"><svg viewBox="0 0 56 56"><circle cx="28" cy="28" r="22" fill="none" stroke="#e3dcc9" stroke-width="6" stroke-dasharray="138.2" stroke-linecap="round" transform="rotate(-90 28 28)"/><circle cx="28" cy="28" r="22" fill="none" id="arcMC" stroke="#a13c2b" stroke-width="6" stroke-dasharray="138.2" stroke-dashoffset="138.2" stroke-linecap="round" transform="rotate(-90 28 28)"/></svg><div class="g-num"><span id="pMC">—</span><span class="g-pct">%ile</span></div></div><div><div class="g-val" id="vMC">—</div><div class="g-med">Med: __MCGDP_MED__%</div><div class="g-zone" id="zMC">—</div></div></div></div>
    <div class="gauge-card"><div class="gauge-title">Yield Gap (EY−Bond)</div><div class="gw"><div class="g-arc"><svg viewBox="0 0 56 56"><circle cx="28" cy="28" r="22" fill="none" stroke="#e3dcc9" stroke-width="6" stroke-dasharray="138.2" stroke-linecap="round" transform="rotate(-90 28 28)"/><circle cx="28" cy="28" r="22" fill="none" id="arcYG" stroke="#2c6e49" stroke-width="6" stroke-dasharray="138.2" stroke-dashoffset="138.2" stroke-linecap="round" transform="rotate(-90 28 28)"/></svg><div class="g-num"><span id="pYG">—</span><span class="g-pct">%ile</span></div></div><div><div class="g-val" id="vYG">—</div><div class="g-med">Med: __YG_MED__%</div><div class="g-zone" id="zYG">—</div></div></div></div>
  </div>

  <!-- SECTION: INDEX LEVELS -->
  <div class="sec">Index Levels — Nifty 50 · Midcap 150 · Smallcap 250</div>
  <div class="g3">
    <div class="cc"><div class="ch"><span class="ct">Nifty 50 Index</span><span class="cv" id="cN">—</span></div><div class="cw"><canvas id="cNifty"></canvas></div></div>
    <div class="cc"><div class="ch"><span class="ct">Nifty Midcap 150 Index</span><span class="cv" id="cMID">—</span></div><div class="cw"><canvas id="cMidchart"></canvas></div></div>
    <div class="cc"><div class="ch"><span class="ct">Nifty Smallcap 250 Index</span><span class="cv" id="cSC">—</span></div><div class="cw"><canvas id="cSCchart"></canvas></div></div>
  </div>

  <!-- SECTION: NIFTY VALUATION -->
  <div class="sec">Nifty 50 Valuation — P/E &amp; Market Cap</div>
  <div class="g2">
    <div class="cc"><div class="ch"><span class="ct">P/E Ratio</span><span class="cv" id="cPE">—</span></div><div class="cw"><canvas id="cPEchart"></canvas></div></div>
    <div class="cc"><div class="ch"><span class="ct">Market Cap (USD Trillion)</span><span class="cv" id="cMCT">—</span></div><div class="cw"><canvas id="cMCTchart"></canvas></div></div>
  </div>

  <!-- SECTION: EARNING YIELD -->
  <div class="sec">Earning Yield — Nifty 50, then Midcap 150 &amp; Smallcap 250</div>
  <div class="g2">
    <div class="cc sp2"><div class="ch"><span class="ct">Nifty 50 Earning Yield %</span><span class="cv" id="cEY">—</span></div><div class="cw"><canvas id="cEYchart"></canvas></div></div>
  </div>
  <div class="g2">
    <div class="cc"><div class="ch"><span class="ct">Midcap 150 Earning Yield %</span><span class="cv" id="cMEY">—</span></div><div class="cw"><canvas id="cMEYchart"></canvas></div></div>
    <div class="cc"><div class="ch"><span class="ct">Smallcap 250 Earning Yield %</span><span class="cv" id="cSEY">—</span></div><div class="cw"><canvas id="cSEYchart"></canvas></div></div>
  </div>

  <!-- SECTION: VALUATION RATIOS -->
  <div class="sec">Valuation Ratios — PREITY · BEER · MC/GDP</div>
  <div class="g3">
    <div class="cc"><div class="ch"><span class="ct">PREITY Ratio (PE × T-Bill)</span><span class="cv" id="cPREITY">—</span></div><div class="cw"><canvas id="cPREITYchart"></canvas></div></div>
    <div class="cc"><div class="ch"><span class="ct">BEER Ratio</span><span class="cv" id="cBEER">—</span></div><div class="cw"><canvas id="cBEERchart"></canvas></div></div>
    <div class="cc"><div class="ch"><span class="ct">MC/GDP % (Buffett)</span><span class="cv" id="cMC">—</span></div><div class="cw"><canvas id="cMCchart"></canvas></div></div>
  </div>

  <!-- SECTION 3: EPS GROWTH -->
  <div class="sec">Earnings — YoY Growth % (trailing 252 sessions ≈ 1 year)</div>
  <div class="g3">
    <div class="cc"><div class="ch"><span class="ct">Nifty 50 EPS Growth % (YoY)</span><span class="cv" id="cNEG">—</span></div><div class="cw"><canvas id="cNEGchart"></canvas></div></div>
    <div class="cc"><div class="ch"><span class="ct">Midcap 150 EPS Growth % (YoY)</span><span class="cv" id="cMEGg">—</span></div><div class="cw"><canvas id="cMEGgchart"></canvas></div></div>
    <div class="cc"><div class="ch"><span class="ct">Smallcap 250 EPS Growth % (YoY)</span><span class="cv" id="cSEGg">—</span></div><div class="cw"><canvas id="cSEGgchart"></canvas></div></div>
  </div>

  <div class="sec">Earnings — Absolute EPS Level (per index unit)</div>
  <div class="g3">
    <div class="cc"><div class="ch"><span class="ct">Nifty 50 EPS Level</span><span class="cv" id="cNEL">—</span></div><div class="cw"><canvas id="cNELchart"></canvas></div></div>
    <div class="cc"><div class="ch"><span class="ct">Midcap 150 EPS Level</span><span class="cv" id="cMEG">—</span></div><div class="cw"><canvas id="cMEGchart"></canvas></div></div>
    <div class="cc"><div class="ch"><span class="ct">Smallcap 250 EPS Level</span><span class="cv" id="cSEG">—</span></div><div class="cw"><canvas id="cSEGchart"></canvas></div></div>
  </div>

  <!-- SECTION 4: BONDS & FX -->
  <div class="sec">Bond Yields &amp; Dollar Index</div>
  <div class="g3">
    <div class="cc sp2"><div class="ch"><span class="ct">India 10yr vs US 10yr &amp; Yield Gap</span><span class="cv" id="cBond">—</span></div><div class="cw"><canvas id="cBondchart"></canvas></div><div class="lgd"><div class="li"><div class="ld" style="background:#9a7321"></div>India 10yr</div><div class="li"><div class="ld" style="background:#1d4634"></div>US 10yr</div><div class="li"><div class="ld" style="background:#2c6e49"></div>Yield Gap</div></div></div>
    <div class="cc"><div class="ch"><span class="ct">USD / INR</span><span class="cv" id="cFX">—</span></div><div class="cw"><canvas id="cFXchart"></canvas></div></div>
  </div>

</div>

<script>
const RAW=__CHART_DATA__;
const STATS=__STATS_DATA__;
const CHAMP=__CHAMP_DATA__;

Chart.defaults.color='#9aa499';
Chart.defaults.font.family="'IBM Plex Mono',monospace";
Chart.defaults.font.size=9;
Chart.defaults.animation={duration:200};

const GRID={color:'rgba(220,211,191,0.7)',lineWidth:1};
const TICK={color:'#9aa499',maxTicksLimit:5};

function bo(){
  return{responsive:true,maintainAspectRatio:false,resizeDelay:50,
    interaction:{mode:'index',intersect:false},
    plugins:{legend:{display:false},tooltip:{backgroundColor:'#fffefb',borderColor:'#dcd3bf',borderWidth:1,titleColor:'#15281e',bodyColor:'#6b7a6f',padding:8}},
    scales:{x:{grid:GRID,ticks:{...TICK,maxRotation:0,maxTicksLimit:6}},y:{grid:GRID,ticks:TICK}}};
}

function grad(id,c){
  const cv=document.getElementById(id);if(!cv)return c+'33';
  const g=cv.getContext('2d').createLinearGradient(0,0,0,130);
  g.addColorStop(0,c+'55');g.addColorStop(1,c+'00');return g;
}

function lds(data,color,fill,id){
  return{data,borderColor:color,backgroundColor:fill?grad(id,color):'transparent',borderWidth:1.5,pointRadius:0,tension:0.3,fill:!!fill};
}

function barDs(data,color){
  return{type:'bar',data,backgroundColor:data.map(v=>v>=0?color+'99':v<0?'#ff5c6a99':color+'99'),borderColor:data.map(v=>v>=0?color:'#a13c2b'),borderWidth:1,borderRadius:2};
}

function lbl(dates){
  return dates.map(d=>new Date(d).toLocaleDateString('en-IN',{day:'2-digit',month:'short',year:'2-digit'}));
}

function pct(arr,v){return Math.round(arr.filter(x=>x<=v).length/arr.length*100);}

function zone(p,flip){
  const cheap=flip?p>60:p<30,rich=flip?p<30:p>70;
  if(cheap)return['Attractive','zc'];if(rich)return['Stretched','zr'];return['Fair Value','zf'];
}

function set(id,v){const el=document.getElementById(id);if(el)el.textContent=v;}

function setGauge(arcId,pId,vId,zId,p,val,flip,color){
  const arc=document.getElementById(arcId);
  arc.style.strokeDashoffset=138.2*(1-p/100);arc.style.stroke=color;
  document.getElementById(pId).textContent=p+'%';
  document.getElementById(vId).textContent=typeof val==='number'?val.toFixed(2):val;
  const[zt,zc]=zone(p,flip);const el=document.getElementById(zId);el.textContent=zt;el.className='g-zone '+zc;
}

function growthColor(v){return v>=0?'var(--green)':'var(--red)';}

const charts={};

function filterData(n){
  const len=RAW.date.length,start=n>=9999?0:Math.max(0,len-n);
  const sl=k=>RAW[k].slice(start);
  return{date:sl('date'),nifty50:sl('nifty50'),midcap150:sl('midcap150'),smallcap250:sl('smallcap250'),pe:sl('pe'),pb:sl('pb'),
    earning_yield:sl('earning_yield'),india_10yr:sl('india_10yr'),us_10yr:sl('us_10yr'),
    yield_gap:sl('yield_gap'),usdinr:sl('usdinr'),dollar_index:sl('dollar_index'),
    marketcap_gdp:sl('marketcap_gdp'),marketcap_trillion:sl('marketcap_trillion'),
    beer:sl('beer'),preity:sl('preity'),
    midcap_earn_yield:sl('midcap_earn_yield'),smallcap_earn_yield:sl('smallcap_earn_yield'),
    eps:sl('eps'),midcap_eps:sl('midcap_eps'),smallcap_eps:sl('smallcap_eps'),
    nifty_eps_growth:sl('nifty_eps_growth'),midcap_eps_growth:sl('midcap_eps_growth'),smallcap_eps_growth:sl('smallcap_eps_growth')};
}

function updateKPIs(d){
  const n=d.date.length-1,p=n>0?n-1:0;
  const nifty=d.nifty50[n],pe=d.pe[n],pb=d.pb[n],ey=d.earning_yield[n];
  const i10=d.india_10yr[n],u10=d.us_10yr[n],yg=d.yield_gap[n];
  const fx=d.usdinr[n],mc=d.marketcap_gdp[n],beer=d.beer[n];
  const mct=d.marketcap_trillion[n],preity=d.preity[n];
  const mey=d.midcap_earn_yield[n],sey=d.smallcap_earn_yield[n];
  const neg=d.nifty_eps_growth[n];
  const chg=((nifty-d.nifty50[p])/d.nifty50[p]*100).toFixed(2);

  set('kN',nifty.toLocaleString('en-IN',{maximumFractionDigits:0}));
  const kNc=document.getElementById('kNc');kNc.textContent=(chg>0?'+':'')+chg+'%';kNc.className='kpi-c '+(chg>=0?'up':'dn');
  set('kPE',pe.toFixed(2));set('kPEm','med __PE_MED__');set('kPB',pb.toFixed(2));set('kEY',ey.toFixed(2)+'%');
  set('kI10',i10.toFixed(2)+'%');
  const kYG=document.getElementById('kYG');kYG.textContent='gap '+(yg>=0?'+':'')+yg.toFixed(2)+'%';kYG.className='kpi-c '+(yg>=0?'up':'dn');
  set('kU10',u10.toFixed(2)+'%');set('kSp','spr '+(i10-u10).toFixed(2)+'%');
  set('kFX',fx.toFixed(2));set('kDXY','DXY '+d.dollar_index[n].toFixed(2));
  const mcEl=document.getElementById('kMC');mcEl.textContent=mc.toFixed(1)+'%';mcEl.className='kpi-v '+(mc>150?'r':mc>120?'gld':'g');
  set('kMCT','$'+mct.toFixed(2)+'T');
  set('kPREITY',preity.toFixed(1));
  set('kMEY',mey.toFixed(2)+'%');set('kSEY',sey.toFixed(2)+'%');
  set('kBEER',beer.toFixed(3));

  const negEl=document.getElementById('kNEG');negEl.textContent=(neg>=0?'+':'')+neg.toFixed(1)+'%';negEl.className='kpi-v '+(neg>=0?'g':'r');
  set('kMEG',d.midcap_eps[n].toFixed(1));
  set('kSEG',d.smallcap_eps[n].toFixed(1));

  set('cN',nifty.toLocaleString('en-IN',{maximumFractionDigits:0}));
  set('cMID',d.midcap150[n].toLocaleString('en-IN',{maximumFractionDigits:0}));
  set('cSC',d.smallcap250[n].toLocaleString('en-IN',{maximumFractionDigits:0}));
  set('cPE',pe.toFixed(2)+'x');set('cEY',ey.toFixed(2)+'%');set('cBEER',beer.toFixed(3));
  set('cMC',mc.toFixed(1)+'%');set('cMCT','$'+mct.toFixed(2)+'T');
  set('cMEY',mey.toFixed(2)+'%');set('cSEY',sey.toFixed(2)+'%');
  set('cPREITY',preity.toFixed(0));
  set('cBond',i10.toFixed(2)+'% / '+u10.toFixed(2)+'%');set('cFX',fx.toFixed(2));
  set('cNEG',(neg>=0?'+':'')+neg.toFixed(1)+'%');
  set('cMEG',d.midcap_eps[n].toFixed(1));
  set('cSEG',d.smallcap_eps[n].toFixed(1));
  set('cNEL',d.eps[n].toFixed(1));
  const meg=d.midcap_eps_growth[n],seg=d.smallcap_eps_growth[n];
  set('cMEGg',(meg>=0?'+':'')+meg.toFixed(1)+'%');
  set('cSEGg',(seg>=0?'+':'')+seg.toFixed(1)+'%');

  const peArr=RAW.pe,beerArr=RAW.beer.filter(x=>x>0),mcArr=RAW.marketcap_gdp.filter(x=>x>0),ygArr=RAW.yield_gap;
  setGauge('arcPE','pPE','vPE','zPE',pct(peArr,pe),pe,false,'#1d4634');
  setGauge('arcBEER','pBEER','vBEER','zBEER',pct(beerArr,beer),beer,true,'#9a7321');
  setGauge('arcMC','pMC','vMC','zMC',pct(mcArr,mc),mc,false,'#a13c2b');
  setGauge('arcYG','pYG','vYG','zYG',pct(ygArr,yg),yg,true,'#2c6e49');

  const score=Math.round((pct(peArr,pe)+(100-pct(beerArr,beer))+pct(mcArr,mc)+(100-pct(ygArr,yg)))/4);
  set('eviNum',score);document.getElementById('eviDot').style.left=score+'%';
  const[hl,desc]=score<35?['Market Attractive','PE, BEER and yield metrics below historical medians. Strong long-term entry zone.']:score<55?['Fairly Valued','Valuations near historical medians. Balanced risk-reward.']:score<75?['Mildly Stretched','Several indicators above median. Prefer quality stocks.']:['Expensive','Valuations in top quartile. Consider reducing equity allocation.'];
  set('eviHl',hl);set('eviDesc',desc);
}

function buildCharts(d){
  const labels=lbl(d.date);
  const ref=(arr,v)=>arr.map(()=>v);
  const opt=bo();

  function mk(key,id,datasets,yOpts){
    if(charts[key]){charts[key].destroy();delete charts[key];}
    const canvas=document.getElementById(id);if(!canvas)return;
    const opts=JSON.parse(JSON.stringify(opt));
    if(yOpts)Object.assign(opts.scales.y,yOpts);
    charts[key]=new Chart(canvas.getContext('2d'),{type:'line',data:{labels,datasets},options:opts});
  }

  function mkBar(key,id,datasets,yOpts){
    if(charts[key]){charts[key].destroy();delete charts[key];}
    const canvas=document.getElementById(id);if(!canvas)return;
    const opts=JSON.parse(JSON.stringify(opt));
    opts.scales.x.grid={display:false};
    if(yOpts)Object.assign(opts.scales.y,yOpts);
    charts[key]=new Chart(canvas.getContext('2d'),{type:'bar',data:{labels,datasets},options:opts});
  }

  mk('nifty','cNifty',[lds(d.nifty50,'#1d4634',true,'cNifty')]);
  mk('mid','cMidchart',[lds(d.midcap150,'#2c6e49',true,'cMidchart')]);
  mk('sc','cSCchart',[lds(d.smallcap250,'#6b4a6e',true,'cSCchart')]);
  mk('pe','cPEchart',[lds(d.pe,'#1d4634',false),{data:ref(d.pe,STATS.pe_median),borderColor:'#9a732166',borderDash:[4,3],borderWidth:1,pointRadius:0,tension:0}],{min:15,max:30});
  mk('ey','cEYchart',[lds(d.earning_yield,'#2c6e49',true,'cEYchart')]);
  mk('beer','cBEERchart',[lds(d.beer,'#9a7321',false),{data:ref(d.beer,1.0),borderColor:'#a13c2b66',borderDash:[4,3],borderWidth:1,pointRadius:0,tension:0},{data:ref(d.beer,STATS.beer_median),borderColor:'#2c6e4966',borderDash:[4,3],borderWidth:1,pointRadius:0,tension:0}],{min:0.4,max:2.0});
  mk('mc','cMCchart',[lds(d.marketcap_gdp,'#a13c2b',false),{data:ref(d.marketcap_gdp,STATS.mcgdp_median),borderColor:'#9a732166',borderDash:[4,3],borderWidth:1,pointRadius:0,tension:0}],{min:80,max:170});
  mk('mct','cMCTchart',[lds(d.marketcap_trillion,'#1d4634',true,'cMCTchart')]);
  mk('mey','cMEYchart',[lds(d.midcap_earn_yield,'#2c6e49',true,'cMEYchart')]);
  mk('sey','cSEYchart',[lds(d.smallcap_earn_yield,'#6b4a6e',true,'cSEYchart')]);
  mk('preity','cPREITYchart',[lds(d.preity,'#a8761f',true,'cPREITYchart')]);
  mk('bond','cBondchart',[lds(d.india_10yr,'#9a7321',false),lds(d.us_10yr,'#1d4634',false),lds(d.yield_gap,'#2c6e49',true,'cBondchart')]);
  mk('fx','cFXchart',[lds(d.usdinr,'#6b4a6e',true,'cFXchart')]);

  // EPS Growth bars — skip zero values (first year has no data)
  const negData=d.nifty_eps_growth.map(v=>v===0?null:v);


  function growthBar(key,id,data,color){
    if(charts[key]){charts[key].destroy();delete charts[key];}
    const canvas=document.getElementById(id);if(!canvas)return;
    const opts=JSON.parse(JSON.stringify(opt));
    opts.scales.x.grid={display:false};
    opts.plugins.tooltip.callbacks={label:ctx=>(ctx.parsed.y>=0?'+':'')+ctx.parsed.y.toFixed(1)+'%'};
    charts[key]=new Chart(canvas.getContext('2d'),{
      type:'bar',
      data:{labels,datasets:[{data,backgroundColor:data.map(v=>v===null?'transparent':v>=0?color+'99':'#ff5c6a99'),borderColor:data.map(v=>v===null?'transparent':v>=0?color:'#a13c2b'),borderWidth:1,borderRadius:2}]},
      options:opts
    });
  }
  growthBar('neg','cNEGchart',negData,'#1d4634');
  const megData=d.midcap_eps_growth.map(v=>v===0?null:v);
  const segData=d.smallcap_eps_growth.map(v=>v===0?null:v);
  growthBar('megg','cMEGgchart',megData,'#2c6e49');
  growthBar('segg','cSEGgchart',segData,'#6b4a6e');
  mk('nel','cNELchart',[lds(d.eps,'#1d4634',true,'cNELchart')]);
  mk('meg','cMEGchart',[lds(d.midcap_eps,'#2c6e49',true,'cMEGchart')]);
  mk('seg','cSEGchart',[lds(d.smallcap_eps,'#6b4a6e',true,'cSEGchart')]);
}

function refresh(range){
  const d=filterData(range);
  updateKPIs(d);
  buildCharts(d);
}

document.getElementById('fr').addEventListener('click',e=>{
  const btn=e.target.closest('[data-r]');if(!btn)return;
  document.querySelectorAll('.fb').forEach(b=>b.classList.remove('on'));
  btn.classList.add('on');refresh(parseInt(btn.dataset.r));
});

function renderChampion(){
  const c=CHAMP;
  const el=document.getElementById('champScore');
  el.textContent=c.score.toFixed(0);
  const col=c.score<40?'var(--green)':c.score>60?'var(--red)':'var(--gold)';
  el.style.background='linear-gradient(135deg,'+col+',var(--accent))';
  el.style.webkitBackgroundClip='text';el.style.webkitTextFillColor='transparent';el.style.backgroundClip='text';
  document.getElementById('champLabel').textContent=c.label+' — '+(c.label==='CHEAP'?'Strong entry zone':c.label==='EXPENSIVE'?'Caution, thin margin of safety':'Balanced risk-reward');
  document.getElementById('champLabel').style.color=col;
  document.getElementById('champSub').innerHTML='P/E percentile <b>'+c.pe_pctile+'</b> · MarketCap/GDP percentile <b>'+c.mc_pctile+'</b> &nbsp;|&nbsp; 0 = cheapest, 100 = most expensive in history';
  const expEl=document.getElementById('champExp');
  expEl.textContent=(c.exp_return>=0?'+':'')+c.exp_return.toFixed(1)+'%';
  expEl.style.color=c.exp_return>=12?'var(--green)':c.exp_return>=6?'var(--gold)':'var(--red)';

  // buckets
  const labels=['Cheapest','Q2','Q3','Q4','Priciest'];
  const bc=document.getElementById('champBuckets');
  bc.innerHTML='';
  c.buckets.forEach((v,i)=>{
    const d=document.createElement('div');
    d.className='cb'+(i===c.bucket?' active':'');
    const vc=v>=12?'var(--green)':v>=6?'var(--gold)':'var(--red)';
    d.innerHTML='<div class="cb-v" style="color:'+vc+'">'+(v>=0?'+':'')+v.toFixed(0)+'%</div><div class="cb-l">'+labels[i]+'</div>';
    bc.appendChild(d);
  });

  // hit rates
  document.getElementById('hitCheapBall').textContent=c.cheap_hit+'%';
  document.getElementById('hitCheapPct').textContent=c.cheap_hit+'%';
  document.getElementById('hitCheapAvg').textContent=(c.cheap_avg>=0?'+':'')+c.cheap_avg+'%';
  document.getElementById('hitExpBall').textContent=c.exp_hit+'%';
  document.getElementById('hitExpPct').textContent=c.exp_hit+'%';
  document.getElementById('hitExpAvg').textContent=(c.exp_avg>=0?'+':'')+c.exp_avg+'%';
}
renderChampion();

let _rzT;
window.addEventListener('resize',function(){
  clearTimeout(_rzT);
  _rzT=setTimeout(function(){
    const on=document.querySelector('.fb.on');
    if(on)refresh(parseInt(on.dataset.r));
  },150);
});

refresh(365);
</script>
</body>
</html>
"""

def main():
    if not EXCEL.exists():
        print(f"ERROR: {EXCEL} not found"); sys.exit(1)
    print(f"📖  Reading {EXCEL.name} …")
    wb   = load_workbook_safe(EXCEL)
    rows = extract_data(wb)
    if not rows:
        print("ERROR: No data rows found"); sys.exit(1)
    print(f"    {len(rows)} rows  |  {rows[0]['date']} → {rows[-1]['date']}")
    cd    = build_chart_data(rows)
    stats = compute_stats(rows)
    champ = compute_champion(rows)
    html = HTML
    html = html.replace("__LAST_DATE__",   stats["last_date"])
    html = html.replace("__DATE_FROM__",   stats["date_from"])
    html = html.replace("__TOTAL_ROWS__",  str(stats["total_rows"]))
    html = html.replace("__PE_MED__",      str(stats["pe_median"]))
    html = html.replace("__BEER_MED__",    str(stats["beer_median"]))
    html = html.replace("__MCGDP_MED__",   str(stats["mcgdp_median"]))
    html = html.replace("__YG_MED__",      str(stats["yg_median"]))
    html = html.replace("__CHART_DATA__",  json.dumps(cd))
    html = html.replace("__STATS_DATA__",  json.dumps(stats))
    html = html.replace("__CHAMP_DATA__",  json.dumps(champ))
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(html, encoding="utf-8")
    print(f"✅  Dashboard → {OUTPUT}  ({OUTPUT.stat().st_size//1024} KB)")
    print(f"    Nifty {stats['nifty']:,.0f}  |  PE {stats['pe']:.2f}  |  MC ${stats['marketcap_trillion']:.2f}T")
    print(f"    Midcap EY {stats['midcap_earn_yield']:.2f}%  |  SC EY {stats['smallcap_earn_yield']:.2f}%")
    print(f"    Champion Signal: {champ['score']:.0f}/100 ({champ['label']})  →  expected 12M return {champ['exp_return']:+.1f}%")
    print(f"    EPS YoY Growth → Nifty {stats['nifty_eps_growth']:+.1f}%  Midcap {stats['midcap_eps_growth']:+.1f}%  SC {stats['smallcap_eps_growth']:+.1f}%")

if __name__ == "__main__":
    main()
